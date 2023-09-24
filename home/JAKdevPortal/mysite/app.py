"""
    Tables available:
    1. 2A
    2. 2B
    3. 3A
    4. 4A
    5. staff
    6. counsellor
    7. leaveletters
    8. general table
    9. materials table
    10. announcement table
    class tables structure
    1.rollno 2.password 3.name 4.boarding 5.cgpa 6.email 7.phoneno 8.certifications 9.projects 10.prizes 11.internships 12.parentNo

    Staff table structure
    1.userid 2.password 3.Name 4.Qualification 5.Designation 6.subjects 7.code 8.papers 9.email 10.classes 11.phoneno

    Counsellors
    1.start 2.end 3.counsellor
    general table
    1.class 2.subjects 3.staffs

    leave table
    1.appliedBy 2.appliedOn 3.fromDate 4.toDate 5.leaveType 6.reason 7.status 8.counsellor

    materials table
    1.subject 2.unit I 3.unit II 4.unit III 5.unit IV 6.unit V

    announcement table
    1.announcement 2.fromDate 3.fromTime 4.toDate 5.toTime 4.class


"""
from flask import Flask,render_template,request,redirect,url_for,flash,session
import sqlite3
from openpyxl import load_workbook  ,Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from datetime import datetime,date,timedelta
import os
import shutil
#import pywhatkit
from urllib.parse import quote, unquote
from flask import flash
import threading
import time
import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
black_border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
)


app = Flask(__name__)
app.secret_key = "super secret key"
app.config['SESSION_PERMANENT'] = False
sqliteConnection = sqlite3.connect('/home/JAKdevPortal/mysite/jak', check_same_thread=False)
cursor = sqliteConnection.cursor()

data=()

def selectQueryHelp( temp):
    l=[]
    for i in temp:
        l.append(i[0])
    return l
def existingClasses():
    cursor.execute("select class from general")
    temp=cursor.fetchall()
    clses=selectQueryHelp(temp)
    return clses


@app.route('/')
def index():
    units=["UNIT I","UNIT II","UNIT III","UNIT IV","UNIT V"]
    session["units"]=units
    announcements=getAnnouncements()
    return render_template('index.html',announcements=announcements)

@app.route('/profile')
def profile():
    if 'logged_in' in session and session['logged_in']:
        return render_template('profile.html',data=session["data"],cls=session["class"],counsellor=session["counsellor"],profile=session["profile"],leave=session["leavePercentage"], user_is_logged_in=True)
    else:
        return render_template('profile.html',data=session["data"],cls=session["class"],counsellor=session["counsellor"],profile=session["profile"],leave=session["leavePercentage"], user_is_logged_in=False)
@app.route('/digitalMaterial')
def digitalMaterial():
    cursor.execute("select subjects from general where class = ?",(session["class"],))
    subjects=cursor.fetchone()[0].split(',')
    session["subjects"]=subjects
    return render_template('digitalMaterial.html',subjects=session["subjects"],units=session["units"],selected="None")

@app.route('/showMaterial',methods=['POST'])
def showMaterial():
    if request.method=='POST':
        subject=request.form.get('subject')
        subjectStaff=cursor.execute("select staffs from general where class = ?",(session["class"],)).fetchone()[0].split(',')[session["subjects"].index(subject)]
        mainMaterials=cursor.execute("select unit1,unit2,unit3,unit4,unit5 from mainmaterials where subject = ? and staff=?", (subject,subjectStaff)).fetchone()
        additionalMaterials=cursor.execute("select unit1,unit2,unit3,unit4,unit5 from additionalmaterials where subject = ?",(subject,)).fetchone()
        try:
            mainMaterials,additionalMaterials=emptyMaterialsRemoverForStudent(mainMaterials,additionalMaterials)
            additionalMaterials=[i.split("%")[:-1] for i in additionalMaterials]
        except:
            mainMaterials=[]
            additionalMaterials=[]
        return render_template('digitalMaterial.html',subjects=session["subjects"],units=session["units"],selected=subject,mainMaterials=mainMaterials,additionalMaterials=additionalMaterials,cls=session["class"])

@app.route('/internalMark')
def internalMark():
    return render_template('internalmark.html',selected="None")

@app.route('/studentInternalMark',methods=['POST'])
def studentInternalMark():
    if request.method=="POST":
        test=request.form.get('test')
        subjects=cursor.execute("select subjects from general where class = ?",(session["class"],)).fetchone()[0].split(',')[:-1]
        path="/home/JAKdevPortal/mysite/static/internalmark/"+session["class"]+"/"
        studentName=session["name"]
        marks=[]
        for subject in subjects:
            path+=subject+".xlsx"
            wb=load_workbook(path)
            sheet=wb.active
            #check if that file has column name test
            updatedOrNot=False
            for i in range(1,sheet.max_column+1):
                if sheet.cell(row=1,column=i).value==test:
                    updatedOrNot=True
                    col=i
                    break
            if updatedOrNot==False:
                marks.append("nil")
            else:
                for i in range(1,sheet.max_row+1):
                    if sheet.cell(row=i,column=2).value==studentName:
                        marks.append(sheet.cell(row=i,column=col).value)
                        break
            path="/home/JAKdevPortal/mysite/static/internalmark/"+session["class"]+"/"

        return render_template('internalmark.html',selected=test,subjects=subjects,marks=marks,cls=session["class"])
@app.route('/leaveApplication')
def leaveApplication():
    cursor.execute("select * from leave where appliedBy = ?",(session["name"]+"("+str(session["rollno"])+")",))
    data=cursor.fetchall()
    return render_template('leaveApplication.html',requests=data)

@app.route('/certifications')
def certifications():
    certificatesExtn=os.listdir("/home/JAKdevPortal/mysite/static/certificates/"+session["class"]+"/"+str(session["rollno"])+"/")
    certificates=[]
    for i in certificatesExtn:
        certificates.append(i.split('.'))
    location="/home/JAKdevPortal/mysite/static/certificates/"+session["class"]+"/"+str(session["rollno"])+"/"
    return render_template('certifications.html',certificates=certificates,location=location)

@app.route('/projects')
def projects():
    projectsExtn=os.listdir("/home/JAKdevPortal/mysite/static/projects/"+session["class"]+"/"+str(session["rollno"])+"/")
    projects=[]
    for i in projectsExtn:
        projects.append(i.split('.'))
    location="/home/JAKdevPortal/mysite/static/projects/"+session["class"]+"/"+str(session["rollno"])+"/"
    return render_template('projects.html',projects=projects,location=location)

@app.route('/prizes')
def prizes():
    prizesExtn=os.listdir("/home/JAKdevPortal/mysite/static/prizes/"+session["class"]+"/"+str(session["rollno"])+"/")
    prizes=[]
    for i in prizesExtn:
        prizes.append(i.split('.'))
    location="/home/JAKdevPortal/mysite/static/prizes/"+session["class"]+"/"+str(session["rollno"])+"/"
    return render_template('prizes.html',prizes=prizes,location=location)

@app.route('/internships')
def internships():
    internshipsExtn=os.listdir("/home/JAKdevPortal/mysite/static/internships/"+session["class"]+"/"+str(session["rollno"])+"/")
    internships=[]
    for i in internshipsExtn:
        internships.append(i.split('.'))
    location="/home/JAKdevPortal/mysite/static/internships/"+session["class"]+"/"+str(session["rollno"])+"/"
    return render_template('internships.html',internships=internships,location=location)

@app.route('/settings')
def settings():
    session["data"]=cursor.execute("select * from {} where rollno = ?".format(session["class"]),(session["rollno"],)).fetchone()
    data=session["data"]
    cls=session["class"]
    return render_template('settings.html',data=data,cls=cls,profile=session["profile"])

@app.route("/updateProfile",methods=['POST'])
def updateProfile():
    if request.method=="POST":
        newProfile=request.files['newProfile']
        phoneno=request.form.get('whatsapp-number')
        parentPhoneNo=request.form.get("parent-whatsapp-number")
        cgpa=request.form.get('cgpa')
        linkedIn=request.form.get('linkedIn')
        try:
            cgpaFile=request.files['cgpaFile']
        except:
            cgpaFile=''
        email=request.form.get('email')

        password=request.form.get('password')
        if request.form["clicked"]=="Update":
            if newProfile.filename!="":
                try:
                    os.remove("/home/JAKdevPortal/mysite/static/profiles/"+session["class"]+"/"+session["profile"])
                except:
                    pass
                extension=newProfile.filename.split('.')[-1]
                newProfile.save("/home/JAKdevPortal/mysite/static/profiles/"+session["class"]+"/"+str(session["rollno"])+"."+extension)
                session["profile"]=str(session["rollno"])+"."+extension
            try:
                if cgpaFile.filename!="":
                    try:
                        os.remove("/home/JAKdevPortal/mysite/static/cgpa/"+session["class"]+"/"+str(session["rollno"]))
                    except:
                        pass
                    extension=cgpaFile.filename.split('.')[-1]
                    cgpaFile.save("/home/JAKdevPortal/mysite/static/cgpa/"+session["class"]+"/"+str(session["rollno"])+"."+extension)
                    cursor.execute("update {} set cgpaFile = ? where rollno = ?".format(session["class"]),(str(session["rollno"])+"."+extension,session["rollno"]))
            except:
                pass
            temp=[phoneno,cgpa,email,password,parentPhoneNo,linkedIn]
            columnName=["phoneno","cgpa","email","password","parentNo","linkedin"]
            for i in range(len(temp)):
                if type(temp[i])==str and temp[i]!="":
                    print(temp[i],columnName[i],session["rollno"])
                    cursor.execute("update {} set {} =? where rollno = ?".format(session["class"],columnName[i]),(temp[i],session["rollno"]))
            sqliteConnection.commit()
            flash('Updated Successfully!!...')
            return redirect(url_for('settings'))


@app.route("/staffUpdateProfile",methods=['POST'])
def staffUpdateProfile():
    if request.method=="POST":
        newProfile=request.files['newProfile']
        phoneno=request.form.get('whatsapp-number')
        designation=request.form.get('designation')
        qualification=request.form.get('qualification')
        email=request.form.get('email')
        username=request.form.get('username')
        password=request.form.get('password')
        if request.form["clicked"]=="Update":
            if newProfile.filename!="":
                try:
                    os.remove("/home/JAKdevPortal/mysite/static/profiles/STAFFS/"+session["profile"])
                except:
                    pass
                extension=newProfile.filename.split('.')[-1]
                compress=session["data"][2].replace(" ","")
                newProfile.save("/home/JAKdevPortal/mysite/static/profiles/STAFFS/"+compress+"."+extension)
                session["profile"]=session["data"][2].replace(" ","")+"."+extension
            temp=[phoneno,email,password,qualification,designation,username]
            columnName=["phoneno","email","password","Qualification","Designation","userid"]
            for i in range(len(temp)):
                if type(temp[i])==str and temp[i]!="":
                    cursor.execute("update staff set {} =? where userid = ?".format(columnName[i]),(temp[i],session["data"][0]))
            sqliteConnection.commit()
            flash('Profile Updated Successfully...')
            return redirect(url_for('staffSettings'))



@app.route('/staffProfile')
def staffProfile():
    data=session["data"]
    papersPublished=len(os.listdir("/home/JAKdevPortal/mysite/static/papers/"+session["data"][2]+"/"))
    return render_template('staffProfile.html',data=data,profile=session["profile"],papersPublished=papersPublished)

def emptyMaterialsRemoverForStudent(M,A):
    m=[]
    a=[]
    for i in range(len(M)):
        if M[i]=="" and A[i]=="":
            pass
        else:
            m.append(M[i])
            a.append(A[i])
    return m,a

def emptyMaterialsRemoverForStaff(m,a):
    index=len(m)-1
    for i in range(index,-1,-1):
        if m[i]=="" and a[i]=="":
            index-=1
        else:
            break
    return m[:index+1],a[:index+1]

@app.route('/staffMaterials')
def staffMaterials():
    cursor.execute("select subjects,classes from staff where userid = ?",(session["data"][0],))
    data=cursor.fetchall()
    subjectsTemp=data[0][0].split(',')[:-1]
    classesTemp=data[0][1].split(',')[:-1]
    subjects=[]
    classes=[]
    for i in range(len(subjectsTemp)):
        if subjects.count(subjectsTemp[i])==0:
            subjects.append(subjectsTemp[i])
            classes.append(classesTemp[i])

    mainMaterials=[]
    otherMaterials=[]
    for i in range(len(subjects)):
        mm,om=cursor.execute("select unit1,unit2,unit3,unit4,unit5 from mainmaterials where subject = ? and staff=?",(subjects[i],session["data"][2])).fetchall(),cursor.execute("select unit1,unit2,unit3,unit4,unit5 from additionalmaterials where subject = ? and staff=?",(subjects[i],session["data"][2])).fetchall()
        mainMaterials.append(mm[0])
        otherMaterials.append(om[0])
    mainMaterials=[list(i) for i in mainMaterials]
    otherMaterials=[list(i) for i in otherMaterials]
    for i in range(len(mainMaterials)):
        tm,ta=emptyMaterialsRemoverForStaff(mainMaterials[i],otherMaterials[i])
        mainMaterials[i]=tm
        otherMaterials[i]=ta

    try:
        deleteMaterial=session["deleteMaterial"]
        selectedDelete=session["selectedDelete"]
    except:
        deleteMaterial="nil"
        selectedDelete="nil"

    return render_template('staffMaterials.html',subjects=subjects,mainmaterials=mainMaterials,additionalMaterials=otherMaterials,classes=classes,deleteMaterial=deleteMaterial,selectedDelete=selectedDelete)

@app.route('/staffAttendance')
def staffAttendance():
    classes=existingClasses()
    existingYears=[]
    for i in classes:
        existingYears.append(clsToYear[i])
    session["existingYears"]=existingYears
    try:
        path=session["attendanceReportPath"]
    except:
        path="None"
    try:
        studentDetails=session["retake"][-1]
    except:
        studentDetails=[]
    return render_template('staffAttendance.html',studentDetails=studentDetails,existingYears=existingYears,selected="None",path=path)

@app.route('/showAttendance',methods=["GET","POST"])
def showStudents():
    if request.method=="POST":
        existingYears=session["existingYears"]
        cls=request.form.get("class")
        yr=yearToCls[cls]
        session["attendanceYr"]=yr
        path="/home/JAKdevPortal/mysite/static/attendance/"+yr+".xlsx"
        if request.form["clicked"]=="Show":
            wb=load_workbook(path)
            sheet=wb.active
            today=date.today()
            for i in range(4,sheet.max_column+1):
                if sheet.cell(row=1,column=i).value==str(today):
                    return render_template('staffAttendance.html',studentDetails="taken",existingYears=session["existingYears"],selected=cls,path=path)

            studentDetails=[]
            cursor.execute("select rollno,name from {}".format(yr))
            studentDetails.append(cursor.fetchall())
            return render_template('staffAttendance.html',studentDetails=studentDetails,existingYears=existingYears,selected=cls,path=path)
        elif request.form["clicked"]=="Retake":
            retakeDate=request.form["retakeDate"]
            if retakeDate=="":
                flash("Please select a date")
                return render_template('staffAttendance.html',studentDetails=[],existingYears=session["existingYears"],selected=cls,path=path)
            session["retakeDate"]=retakeDate
            studentDetails=[]
            cursor.execute("select rollno,name from {}".format(yr))
            studentDetails.append(cursor.fetchall())
            print(studentDetails)
            return render_template('staffAttendance.html',studentDetails=studentDetails,existingYears=existingYears,selected=cls,path=path)

@app.route('/uploadAttendance',methods=["GET","POST"])
def uploadAttendance():
    if request.method=="POST":
        attendanceYr=session["attendanceYr"]
        path="/home/JAKdevPortal/mysite/static/attendance/"+attendanceYr+".xlsx"
        wb=load_workbook(path)
        sheet=wb.active
        print(attendanceYr)
        regnos=cursor.execute("select rollno from {}".format(attendanceYr)).fetchall()
        present=[]
        for i in regnos:
            present.append(request.form.get(str(i[0])+"_attendance"))
        if request.form['clicked']=="Update":
            try:
                retakeDate=session["retakeDate"]
            except:
                retakeDate=""
            if retakeDate=="":
                col=sheet.max_column+1
                print(col)
                date=str(datetime.date(datetime.now()))
                sheet.cell(row=1,column=col).value=date
                for i in range(2,sheet.max_row+1):
                    if present[i-2]=='A':
                        sheet.cell(row=i,column=col).value="AB"
                    elif present[i-2]=='OD':
                        sheet.cell(row=i,column=col).value="OD"

                wb.save(path)
                flash("Attendance Updated Successfully")
                return render_template('staffAttendance.html',studentDetails=[],existingYears=session["existingYears"],selected="None",path=path)


            else:
                session["retakeDate"]=""
                presentFlag=False
                for i in range(4,sheet.max_column+1):
                    #change the format of retake date to %m-%d-%Y
                    print(retakeDate)
                    if sheet.cell(row=1,column=i).value==retakeDate:
                        presentFlag=True
                        col=i
                        break
                if presentFlag:
                    for i in range(2,sheet.max_row+1):
                        if present[i-2]=="A":
                            sheet.cell(row=i,column=col).value="AB"
                        elif present[i-2]=="OD":
                            sheet.cell(row=i,column=col).value="OD"
                        else:
                            sheet.cell(row=i,column=col).value=""
                    wb.save(path)
                else:
                    col=sheet.max_column+1
                    date=retakeDate.format("%d-%m-%Y")
                    sheet.cell(row=1,column=col).value=date
                    for i in range(2,sheet.max_row+1):
                        if present[i-2]=="A":
                            sheet.cell(row=i,column=col).value="AB"
                        elif present[i-2]=="OD":
                            sheet.cell(row=i,column=col).value="OD"
                        else:
                            sheet.cell(row=i,column=col).value=""
                    wb.save(path)
                #sort the value of the sheet as per date in ascending order along with its column values
                #sort the value of the sheet as per date in ascending order along with its column values
                dates_and_values = []
                for i in range(4, sheet.max_column + 1):
                    col_date = sheet.cell(row=1, column=i).value
                    col_values = [sheet.cell(row=j, column=i).value for j in range(2, sheet.max_row + 1)]
                    dates_and_values.append((col_date, col_values))

                # Sort the column headers (dates) based on date values
                dates_and_values.sort(key=lambda x: datetime.strptime(x[0], "%Y-%m-%d"))

                # Rearrange columns based on the sorted dates
                for i, (col_date, col_values) in enumerate(dates_and_values):
                    # Update the column header with the sorted date
                    sheet.cell(row=1, column=i + 4).value = col_date
                    # Update the corresponding column values
                    for j, value in enumerate(col_values):
                        sheet.cell(row=j + 2, column=i + 4).value = value

                # Save the workbook
                wb.save(path)


                flash("Attendance Updated Successfully")
                return render_template('staffAttendance.html',studentDetails=[],existingYears=session["existingYears"],selected="None",path=path)

        elif request.form['clicked']=="Upload":
            attendanceFile=request.files["attendanceFile"]
            os.remove(path)
            attendanceFile.save(path)
            flash('Uploaded Successfully...')
            return render_template('staffAttendance.html',studentDetails=[],existingYears=session["existingYears"],selected="None",path=path)

def generateExcelForMonthlyAttendance(month,cls):
    os.mkdir("/home/JAKdevPortal/mysite/static/temp")
    clsD={"A2":"II Year Attendance","A3":"III Year Attendance"}
    monthD={"01":"January","02":"February","03":"March","04":"April","05":"May","06":"June","07":"July","08":"August",
            "09":"September","10":"October","11":"November","12":"December"}
    path="/home/JAKdevPortal/mysite/static/temp/"+clsD[cls]+" "+monthD[month]+".xlsx"
    session["attendanceReportPath"]=path
    wb=Workbook()
    sheet=wb.active
    bold_font = Font(bold=True)
    sheet.cell(row=1, column=1).value = "R.M.D. ENGINEERING COLLEGE"
    sheet.cell(row=1, column=1).font = bold_font
    sheet.cell(row=2, column=1).value = "DEPARTMENT OF ARTIFICIAL INTELLIGENCE AND MACHINE LEARNING"
    sheet.cell(row=2, column=1).font = bold_font
    sheet.cell(row=4,column=2).value="S.NO"
    sheet.cell(row=4,column=2).font = bold_font
    sheet.cell(row=4,column=3).value="REG NO"
    sheet.cell(row=4,column=3).font = bold_font
    sheet.cell(row=4,column=4).value="NAME"
    sheet.cell(row=4,column=4).font = bold_font
    #iterate through static/attendance/cls.xlsx find dates which are month
    #if date is in month then add that date to the sheet
    path1="/home/JAKdevPortal/mysite/static/attendance/"+cls+".xlsx"
    wb1=load_workbook(path1)
    sheet1=wb1.active
    total_days=0
    for i in range(4,sheet1.max_column+1):
        print(sheet1.cell(row=1,column=i).value)
        if sheet1.cell(row=1,column=i).value.split('-')[1]==month:
            total_days+=1
            y,m,d=sheet1.cell(row=1,column=i).value.split('-')
            date=d+"-"+m+"-"+y
            sheet.cell(row=4,column=i+1).value=date
            sheet.cell(row=4,column=i+1).font = bold_font

    sheet.cell(row=3, column=2).value = "NO.OF WORKING DAYS : "+str(total_days)
    sheet.cell(row=3, column=2).font = bold_font

    #also add DAYS ABSENT AND ATTENDANCE PERCENTAGE
    sheet.cell(row=4,column=total_days+5).value="DAYS ABSENT"
    sheet.cell(row=4,column=total_days+5).font = bold_font
    sheet.cell(row=4,column=total_days+6).value="ATTENDANCE PERCENTAGE"
    sheet.cell(row=4,column=total_days+6).font = bold_font

    #write students names from
    if cls=="A2":
        #iterate through a2.xlsx and b2.xlsx and write names
        currentRowForP=5
        s_no=1
        for c in ("A2","B2"):
            path2="/home/JAKdevPortal/mysite/static/attendance/"+c+".xlsx"
            wb2=load_workbook(path2)
            sheet2=wb2.active

            #write in the sheet the attendance of each student of given month
            for i in range(2,sheet1.max_row+1):
                sheet.cell(row=currentRowForP,column=2).value=s_no
                sheet.cell(row=currentRowForP,column=3).value=sheet2.cell(row=i,column=1).value
                sheet.cell(row=currentRowForP,column=4).value=sheet2.cell(row=i,column=2).value
                currentRowForP+=1
                s_no+=1

        currentRow=5
        counter=1
        for c in ("A2","B2"):
            path2="/home/JAKdevPortal/mysite/static/attendance/"+c+".xlsx"
            wb2=load_workbook(path2)
            sheet2=wb2.active
            for i in range(4,sheet2.max_column+1):
                if sheet2.cell(row=1,column=i).value.split('-')[1]==month:
                    for j in range(2,sheet2.max_row+1):
                        try:
                            if sheet2.cell(row=j,column=i).value.lower()=='a' or sheet2.cell(row=j,column=i).value.lower()=='ab' :
                                sheet.cell(row=currentRow,column=i+1).value=sheet2.cell(row=j,column=i).value
                        except:
                            pass
                        currentRow+=1
                    if counter==1:
                        currentRow=5
                    else:
                        currentRow=sheet2.max_row+4
            counter+=1
            currentRow=sheet2.max_row+4
            wb2.close()

        #assign days absent and attendance percentage
        currentRow=5
    else:
        currentRow=5
        currentRowForP=5
        s_no=1
        path2="/home/JAKdevPortal/mysite/static/attendance/"+cls+".xlsx"
        wb2=load_workbook(path2)
        sheet2=wb2.active
        for i in range(2,sheet1.max_row+1):
            sheet.cell(row=currentRowForP,column=2).value=s_no
            sheet.cell(row=currentRowForP,column=3).value=sheet2.cell(row=i,column=1).value
            sheet.cell(row=currentRowForP,column=4).value=sheet2.cell(row=i,column=2).value
            currentRowForP+=1
            s_no+=1
        for i in range(4,sheet1.max_column+1):
            if sheet1.cell(row=1,column=i).value.split('-')[1]==month:
                for j in range(2,sheet2.max_row+1):
                    sheet.cell(row=currentRow,column=i+1).value=sheet2.cell(row=j,column=i).value
                    currentRow+=1
            currentRow=5

    for i in range(5,sheet.max_row+1):
        daysAbsent=0
        for j in range(5,sheet.max_column-1):
            try:
                if sheet.cell(row=i,column=j).value.lower()=="a" or sheet.cell(row=i,column=j).value.lower()=="ab" :
                    daysAbsent+=1
            except:
                pass
        sheet.cell(row=i,column=sheet.max_column-1).value=daysAbsent
        try:
            sheet.cell(row=i,column=sheet.max_column).value=str(round((1-daysAbsent/total_days)*100,2))
        except:
            sheet.cell(row=i,column=sheet.max_column).value=0


    for i in range(4,sheet.max_row+1):
        for j in range(2,sheet.max_column+1):
            sheet.cell(row=i,column=j).border=black_border


    wb1.close()
    wb.save(path)


def generateExcelForDailyAttendance(date,cls):
    os.mkdir("/home/JAKdevPortal/mysite/static/temp")
    clsD={"A2":"II Year Attendance","A3":"III Year Attendance"}
    path="/home/JAKdevPortal/mysite/static/temp/"+"AIML Absentees on "+date+".xlsx"
    session["attendanceReportPath"]=path
    wb=Workbook()
    sheet=wb.active
    bold_font = Font(bold=True)
    details=["R.M.D. ENGINEERING COLLEGE","R.S.M. NAGAR, KAVARAIPETTAI","DEPARTMENT OF ARTIFICIAL INTELLIGENCE AND MACHINE LEARNING","DAILY ABSENTEES REPORT","ODD SEMESTER 2023-2024 "]
    for i in range(len(details)):
        sheet.cell(row=i+1,column=1).value=details[i]
        sheet.cell(row=i+1,column=1).font = bold_font
    sheet.cell(row=7,column=1).value="DATE : "
    sheet.cell(row=7,column=1).font = bold_font
    #convert the format of date to %d-%m-%Y
    y,m,d=date.split('-')
    tempDate=d+"-"+m+"-"+y
    sheet.cell(row=7,column=2).value=tempDate
    sheet.cell(row=7,column=2).font = bold_font
    columnNames=["S.No","Register Number","Name","DEPT","YEAR","SEC","M/F","D/H","Reason"]
    for i in range(len(columnNames)):
        sheet.cell(row=8,column=i+1).value=columnNames[i]
        sheet.cell(row=8,column=i+1).font = bold_font
    #save for now
    s_no=1
    currentRow=9
    yr2=0
    yr3=0
    hostellers2yr=0
    dayscholars2yr=0
    hostellers3yr=0
    dayscholars3yr=0
    od2yr=0
    od3yr=0

    for i in ("A2","B2","A3"):
        path1="/home/JAKdevPortal/mysite/static/attendance/"+i+".xlsx"
        wb1=load_workbook(path1)
        sheet1=wb1.active
        for j in range(4,sheet1.max_column+1):
            if sheet1.cell(row=1,column=j).value==date:
                for k in range(2,sheet1.max_row+1):
                    try:
                        if sheet1.cell(row=k,column=j).value.lower()=="ab" or sheet1.cell(row=k,column=j).value.lower()=="od":
                            if sheet1.cell(row=k,column=j).value.lower()=="od":
                                if i=="A2" or i=="B2":
                                    od2yr+=1
                                else:
                                    od3yr+=1


                            sheet.cell(row=currentRow,column=1).value=s_no
                            sheet.cell(row=currentRow,column=2).value=sheet1.cell(row=k,column=1).value
                            sheet.cell(row=currentRow,column=3).value=sheet1.cell(row=k,column=2).value
                            sheet.cell(row=currentRow,column=4).value="AIML"
                            if i=="A2" or i=="B2":
                                sheet.cell(row=currentRow,column=5).value="II"
                            else:
                                sheet.cell(row=currentRow,column=5).value="III"
                            if i=="A2":
                                sheet.cell(row=currentRow,column=6).value="A"
                                yr2+=1
                            elif i=="B2":
                                sheet.cell(row=currentRow,column=6).value="B"
                                yr2+=1
                            else:
                                sheet.cell(row=currentRow,column=6).value=""
                                yr3+=1
                            sex,boarding=cursor.execute("select sex,boarding from {} where rollno = ?".format(i),(sheet1.cell(row=k,column=1).value,)).fetchone()
                            print(sex)
                            print(boarding)
                            if boarding=="Hosteller":
                                print(i)
                                if i=="A2" or i=="B2":
                                    hostellers2yr+=1
                                else:
                                    hostellers3yr+=1
                            else:
                                if i=="A2" or i=="B2":
                                    dayscholars2yr+=1
                                else:
                                    dayscholars3yr+=1
                            sheet.cell(row=currentRow,column=7).value=sex
                            sheet.cell(row=currentRow,column=8).value=boarding[0].upper()
                            reason=cursor.execute("select reason from leave where appliedBy like ? and fromDate = ?",("%"+sheet1.cell(row=k,column=2).value+"%",date)).fetchone()
                            if reason==None:
                                if sheet1.cell(row=k,column=j).value.lower()=="od":
                                    sheet.cell(row=currentRow,column=9).value="OD"
                            else:
                                sheet.cell(row=currentRow,column=9).value=reason[0]

                            currentRow+=1
                            s_no+=1
                    except:
                        pass
        #adding black border to the table
        for i in range(8,currentRow):
            for j in range(1,10):
                cell = sheet.cell(row=i, column=j)
                cell.border = black_border
        wb1.close()

    details=["II Year","III Year","TOTAL"]
    for i in range(len(details)):
        sheet.cell(row=currentRow+3,column=i+4).value=details[i]
        sheet.cell(row=currentRow+3,column=i+4).font = bold_font
    currentRow+=1


    types=["DayScholars","Hostellers","OD"]
    for i in range(len(types)):
        sheet.cell(row=currentRow+3+i,column=3).value=types[i]
        sheet.cell(row=currentRow+3+i,column=3).font = bold_font

    sheet.cell(row=currentRow+3,column=4).value=dayscholars2yr
    sheet.cell(row=currentRow+3,column=5).value=dayscholars3yr
    sheet.cell(row=currentRow+3,column=6).value=dayscholars2yr+dayscholars3yr
    sheet.cell(row=currentRow+4,column=4).value=hostellers2yr
    sheet.cell(row=currentRow+4,column=5).value=hostellers3yr
    sheet.cell(row=currentRow+4,column=6).value=hostellers2yr+hostellers3yr
    sheet.cell(row=currentRow+5,column=4).value=od2yr
    sheet.cell(row=currentRow+5,column=5).value=od3yr
    sheet.cell(row=currentRow+5,column=6).value=od2yr+od3yr
    for i in range(1,6):
        sheet.merge_cells('A'+str(i)+':I'+str(i))
        #cell allignment
        cell = sheet.cell(row=i, column=1)
        cell.alignment = Alignment(horizontal='center', vertical='center')


    wb.save(path)

@app.route('/attendanceReport',methods=["GET","POST"])
def attendanceReport():
    if request.method=="POST":
        attendanceYr=request.form.get("class")
        typeOfReport=request.form.get("reportType")
        month=request.form.get("monthSelector")
        date=request.form.get("dateSelector")
        if request.form["clicked"]=="Generate":
            #remove the directory temp and files inside it
            try:
                shutil.rmtree("/home/JAKdevPortal/mysite/static/temp")
            except:
                pass
            if typeOfReport=="monthly":
                generateExcelForMonthlyAttendance(month,attendanceYr)
                return redirect(url_for("staffAttendance"))
            else:
                generateExcelForDailyAttendance(date ,attendanceYr)
                return redirect(url_for("staffAttendance"))
        return render_template('staffattendance.html',attendanceYr=attendanceYr,typeOfReport=typeOfReport)

@app.route('/staffInternalMarks')
def staffInternalMarks():
    #classes handled by this teacher
    classes,subjects=cursor.execute("select classes,subjects from staff where userid = ?",(session["data"][0],)).fetchall()[0]
    classes=[clsToYear[i] for i in classes.split(',')[:-1]]
    subjects=list(set(subjects.split(',')[:-1]))
    session["intSub"]=subjects
    try:
        selectedCls,selectedSub,selectedTest=session["internals"]
    except:
        selectedCls,selectedSub,selectedTest="None","None","None"
    return render_template('staffInternalMarks.html',classes=classes,subjects=subjects,students=[],selectedClass=selectedCls,selectedSubject=selectedSub,selectedTest=selectedTest)

def getColumnNames(sheet):
    columnName=[]
    for i in range(1,sheet.max_column+1):
        columnName.append(sheet.cell(row=1,column=i).value)
    return columnName

@app.route("/showInternalMarks",methods=["POST"])
def showInternalMarks():
    if request.method=="POST":
        cls=request.form.get("class")
        subject=request.form.get("subject")
        test=request.form.get("test")
        session["internals"]=[cls,subject,test]
        if request.form["clicked"]=="Show":
            path="/home/JAKdevPortal/mysite/static/internalmark/"+yearToCls[cls]+"/"+subject+".xlsx"
            #if any column has the same test name get all those value in a list
            try:
                wb=load_workbook(path)
            except:
                return render_template("staffInternalMarks.html",marks="nil",students=[],selectedClass=cls,selectedSubject=subject,selectedTest=test,classes=[clsToYear[i] for i in existingClasses()],subjects=session["intSub"])
            sheet=wb.active
            marks=[]
            columnNames=getColumnNames(sheet)
            studentNameList=[]
            for i in range(2,sheet.max_row+1):
                studentNameList.append(sheet.cell(row=i,column=2).value)
            if test in columnNames:
                columnIndex=columnNames.index(test)+1
                for i in range (2,sheet.max_row+1):
                    marks.append(sheet.cell(row=i,column=columnIndex).value)
                return render_template("staffInternalMarks.html",marks=marks,students=studentNameList,selectedClass=cls,selectedSubject=subject,selectedTest=test,classes=[clsToYear[i] for i in existingClasses()],subjects=session["intSub"])
            else:
                return render_template("staffInternalMarks.html",marks="nil",students=studentNameList,selectedClass=cls,selectedSubject=subject,selectedTest=test,classes=[clsToYear[i] for i in existingClasses()],subjects=session["intSub"])


@app.route("/getInternalMarks",methods=["POST"])
def getInternalMarks():
    if request.method=="POST":
        internals=session["internals"]
        cls=internals[0]
        subject=internals[1]
        test=internals[2]
        path="/home/JAKdevPortal/mysite/static/internalmark/"+yearToCls[cls]+"/"+subject+".xlsx"
        wb=load_workbook(path)
        sheet=wb.active
        #add new column with name of test
        index=sheet.max_column+1
        if request.form["clicked"]=="Save":
            sheet.cell(row=1,column=index).value=test
            for i in range(2,sheet.max_row+1):
                mark=request.form[sheet.cell(row=i,column=2).value]
                sheet.cell(row=i,column=index).value=mark
                wb.save(path)
            return redirect(url_for("staffInternalMarks"))


@app.route("/uploadInternalMarks",methods=["POST"])
def uploadInternalMarks():
    if request.method=='POST':
        file=request.files["internalMarksFile"]
        cls,subject,test=session["internals"]
        if request.form['clicked']=='Upload':
            file.save("/home/JAKdevPortal/mysite/static/internalmark/"+yearToCls[cls]+"/"+subject+".xlsx")
            return redirect(url_for("staffInternalMarks"))



@app.route('/staffLeaveLetters')
def staffLeaveLetters():
    cursor.execute("select * from leave where counsellor like'%"+session["data"][2].split()[0]+"%' and status = 'pending'")
    data=cursor.fetchall()
    return render_template('staffLeaveLetters.html',requests=data)

@app.route('/leaveHistory')
def leaveHistory():
    try:
        leaves,selectedClass,selectedTime=session["leaveHistory"]
    except:
        leaves=cursor.execute("select * from leave where status != 'pending'").fetchall()
        selectedClass="all"
        selectedTime="all"
    return render_template('leaveHistory.html',leaves=leaves,selectedClass=selectedClass,selectedTime=selectedTime)

@app.route('/leaveHistoryDisplay',methods=['POST'])
def leaveHistoryDisplay():
    if request.method=='POST':
        cls=request.form.get('class')
        time=request.form.get('time')
        leaveD={"3":"A3","4":"A4"}

        if time=="all" and cls=="all":
            leaves=cursor.execute("select * from leave where status != 'pending' and status !='rejected'").fetchall()
        elif cls=="all":
            currentDate=datetime.now()
            fromDate=currentDate-timedelta(days=int(time))
            leaves=cursor.execute("select * from leave where status != 'pending' and status !='rejected' and fromDate >=? and fromDate<=?",(fromDate,currentDate)).fetchall()

        elif time=="all":
            if cls=='2':
                leaves=cursor.execute("select * from leave where status != 'pending' and status !='rejected' and class = ? or class=?",("A2","B2")).fetchall()
            else:
                leaves=cursor.execute("select * from leave where status != 'pending' and status !='rejected' and class = ?",(leaveD[cls],)).fetchall()
        else:
            currentDate=datetime.now()
            fromDate=currentDate-timedelta(days=int(time))
            if cls=='2':
                leaves=cursor.execute("select * from leave where status != 'pending' and status !='rejected' and class = ? or class=? and fromDate >=? and fromDate<=?",("A2","B2",fromDate,currentDate)).fetchall()
            else:
                leaves=cursor.execute("select * from leave where status != 'pending' and status !='rejected' and class = ? and fromDate >=? and fromDate<=?",(leaveD[cls],fromDate,currentDate)).fetchall()
        session["leaveHistory"]=[leaves,cls,time]
        if request.form['clicked']=='Show':
            return redirect(url_for('leaveHistory'))

def check_date_in_range(from_date, to_date, date_to_check):
    from_date = datetime.strptime(from_date, "%Y-%m-%d")
    to_date = datetime.strptime(to_date, "%Y-%m-%d")
    date_to_check = datetime.strptime(date_to_check, "%Y-%m-%d")
    if from_date <= date_to_check <= to_date:
        return True
    else:
        return False

def check_date_exceeds(to_date,date_to_check):
    to_date = datetime.strptime(to_date, "%Y-%m-%d")
    date_to_check = datetime.strptime(date_to_check, "%Y-%m-%d")
    if date_to_check>to_date:
        return True
    else:
        return False

def getAnnouncements():
    today=date.today()
    cursor.execute("select * from announcements")
    data=cursor.fetchall()
    announcements=[]
    for i in data:
        if check_date_in_range(i[1],i[2],str(today)):
            announcements.append(i[0])
        if check_date_exceeds(i[2],str(today)):
            cursor.execute("delete from announcements where announcement=?",(i[0],))
            sqliteConnection.commit()
    return announcements


@app.route('/announcements')
def announcements():
    announcements=selectQueryHelp(cursor.execute("select announcement from announcements").fetchall())
    return render_template('announcements.html',announcements=announcements)

@app.route("/addAnnouncement", methods=["POST"])
def addAnnouncement():
    if request.method=='POST':
        announcement=request.form['announcement']
        fromDate=request.form['fromDate']
        toDate=request.form['toDate']
        cursor.execute("insert into announcements values(?,?,?)",(announcement,fromDate,toDate))
        sqliteConnection.commit()
        flash('Announcement added successfully...')
        return redirect(url_for('announcements'))


@app.route('/deleteAnnouncement/<path:encoded_announcement>')
def deleteAnnouncement(encoded_announcement):
    announcement = unquote(encoded_announcement)
    cursor.execute("DELETE FROM announcements WHERE announcement=?", (announcement,))
    sqliteConnection.commit()
    flash('Announcement deleted...')
    return redirect(url_for('announcements'))

@app.route("/papers")
def papers():
    path="/home/JAKdevPortal/mysite/static/papers/"+session["data"][2]+"/"
    files=os.listdir(path)
    return render_template("papers.html",files=files,location=path)

@app.route("/addPaper",methods=["POST"])
def appPaper():
    if request.method=="POST":
        path="/home/JAKdevPortal/mysite/static/papers/"+session["data"][2]+"/"
        paperName=request.form["paperName"]
        file=request.files["paperFile"]
        extension=file.filename.split(".")[-1]
        if request.form["clicked"]=="Add":
            existingPaper=cursor.execute("select papers from staff where userid = ?",(session["data"][0],)).fetchone()[0]
            existingPaper=existingPaper+paperName+"."+extension+"%"
            cursor.execute("update staff set papers = ? where userid = ?",(existingPaper,session["data"][0]))
            sqliteConnection.commit()
            file.save("/home/JAKdevPortal/mysite/static/papers/"+session["data"][2]+"/"+paperName+"."+extension)
            flash('Papers uploaded successfully')
            return redirect(url_for('papers'))

@app.route("/deletePaper/<location>/<paperName>")
def deletePaper(location,paperName):
    existingPapers=cursor.execute("select papers from staff where userid = ?",(session["data"][0],)).fetchone()[0].split("%")
    existingPapers.remove(paperName)
    cursor.execute("update staff set papers=? where userid=?",(("%").join(existingPapers),session["data"][0]))
    sqliteConnection.commit()
    location=location.replace(",","/")
    os.remove(location+"/"+paperName)
    flash('Papers deleted successfully')
    return redirect(url_for('papers'))


@app.route('/searchStudent')
def searchStudent():
    Classes=existingClasses()
    totalStudentsData=[]
    for i in Classes:
        cursor.execute("select * from "+i)
        temp=cursor.fetchall()
        totalStudentsData.append(temp)
    return render_template('searchStudent.html',classes=[clsToYear[i] for i in Classes],totalStudentsData=totalStudentsData)


@app.route('/studentDetails/<rollno>/<cls>')
def studentDetails(rollno,cls):
    cls=yearToCls[cls]
    cursor.execute("select * from "+cls+" where rollno = ?",(rollno,))
    data=[]
    d=cursor.fetchone()
    for i in d:
        data.append(i)
    cursor.execute("select counsellor from counsellors where ? between start and end",(data[0],))
    try:
        counsellor=cursor.fetchone()[0]
    except:
        counsellor=""


    data.append(counsellor)
    data.append(os.listdir("/home/JAKdevPortal/mysite/static/certificates/"+cls+"/"+str(rollno)+"/"))
    data.append( os.listdir("/home/JAKdevPortal/mysite/static/projects/"+cls+"/"+str(rollno)+"/"))
    data.append(os.listdir("/home/JAKdevPortal/mysite/static/prizes/"+cls+"/"+str(rollno)+"/"))
    data.append(os.listdir("/home/JAKdevPortal/mysite/static/internships/"+cls+"/"+str(rollno)+"/"))
    data.append(cls)
    session["leavePercentage"]=calculateLeavePercentage(data[0],cls)
    x="/home/JAKdevPortal/mysite/static/cgpa/"+cls+"/"
    cgpaFilePath=os.listdir(x)
    for i in cgpaFilePath:
        if i.split(".")[0]==str(rollno):
            data.append(x+i)
            break
    return render_template('studentDetails.html',data=data,cls=cls,leave=session["leavePercentage"])
@app.route("/staffSettings")
def staffSettings():
    session["data"]=cursor.execute("select * from staff where userid=?",(session["data"][0],)).fetchone()
    return render_template("staffSettings.html",data=session["data"],profile=session["profile"])


#ADMIN STUFFS
classes=selectQueryHelp(cursor.execute("select class from general").fetchall())
possibleClasses=["A2","A3","A4","B2"]
clsToYear={"A2":"2nd Year A","A3":"3rd Year A","A4":"4th Year A","B2":"2nd Year B"}
yearToCls={"2nd Year A":"A2","3rd Year A":"A3","4th Year A":"A4","2nd Year B":"B2"}
units={"UNIT I":"unit1","UNIT II":"unit2","UNIT III":"unit3","UNIT IV":"unit4","UNIT V":"unit5"}

@app.route('/createClass')
def createClass():
    clses=existingClasses()
    newClasses=[]
    for i in possibleClasses:
        if i not in clses:
            newClasses.append(clsToYear[i])
    return render_template('createClass.html',classes=newClasses,existingClasses=clses)



@app.route('/newClass',methods=['POST'])
def newClass():
    if request.method=='POST':
        yr=request.form.get('class')
        excelFile=request.files['excelFile']
        cls=yearToCls[yr]
        if request.form["clicked"]=="Create":
            #MAKING NECESSARY DIRECTORIES
            directories=["certificates","internships","prizes","projects","internalmark","profiles","materials","cgpa"]
            for i in directories:
                os.mkdir("/home/JAKdevPortal/mysite/static/"+i+"/"+cls)
            #create an excel file with class name in static/attendance folder using openpyxl for writing attendance
            cursor.execute("insert into general values(?,?,?)",(cls,"",""))
            cursor.execute("CREATE TABLE {}(rollno varchar(12),password varchar(22),name varchar(50),boarding varchar(3),cgpa varchar(4),email varchar(40),phoneno varchar(11),certifications varchar(250),projects varchar(200),prizes varchar(250),internships varchar(300),parentNo varchar(12),linkedin varchar(100),cgpaFile varchar(55),sex varchar(2));".format(cls))
            sqliteConnection.commit()
            excelFile.save("/home/JAKdevPortal/mysite/static/namelist/"+cls+"."+excelFile.filename.split(".")[-1])
            path = "/home/JAKdevPortal/mysite/static/namelist/"+cls+"."+excelFile.filename.split(".")[-1]
            attendance_path="/home/JAKdevPortal/mysite/static/attendance/"+cls+".xlsx"
            internalMark_path="/home/JAKdevPortal/mysite/static/internalmark/"+cls+".xlsx"
            attendanceWB=Workbook()
            internalMarkWB=Workbook()
            sheet=attendanceWB.active
            sheet1=internalMarkWB.active
            columnName=["REG NO","NAME","PHONE NO"]
            sheet.append(columnName)
            sheet1.append(columnName)
            wb=load_workbook(path)
            sheet_obj = wb.active
            m_row = sheet_obj.max_row
            m_col=sheet_obj.max_column
            for i in range(2, m_row + 1):
                studentData=[]
                for j in range(1,m_col+1):
                    cell_obj = sheet_obj.cell(row = i, column = j)
                    studentData.append(cell_obj.value)
                print(studentData)
                cursor.execute("insert into {} values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)".format(cls),(studentData[0],studentData[0],studentData[1],studentData[2],"",studentData[4],studentData[3],"","","","","","","",studentData[5]))
                sheet.append([studentData[0],studentData[1],studentData[3]])
                sheet1.append([studentData[0],studentData[1],studentData[3]])
                for i in directories[:-4]:
                    os.mkdir("/home/JAKdevPortal/mysite/static/"+i+"/"+cls+"/"+str(studentData[0]))
                sqliteConnection.commit()
            attendanceWB.save(attendance_path)
            internalMarkWB.save(internalMark_path)
            flash('Class created successfully...')
            return redirect(url_for('createClass'))

def promoter(classes):
    tests=["Internal Assessment I","Internal Assessment II","Model Exam"]
    for cls in classes:
        for test in tests:
            path="/home/JAKdevPortal/mysite/static/"+cls+test+".xlsx"
            wb=Workbook()
            sheet=wb.active
            sheet.cell(row=1,column=1).value="REG NO"
            sheet.cell(row=1,column=2).value="NAME"
            sheet.cell(row=1,column=3).value="PHONE NO"
            studentMark=internalMarksAdmin(cls,test)
            for sub in range( len(studentMark[0])   ):
                sheet.cell(row=1,column=4+sub).value=studentMark[0][sub]
            for i in range(1,len(studentMark)):
                sheet.cell(row=i+1,column=1).value=studentMark[i][0]
                sheet.cell(row=i+1,column=2).value=studentMark[i][1]
                sheet.cell(row=i+1,column=3).value=studentMark[i][2]
                #subjects Marks
                try:
                    sheet.cell(row=i+1,column=4).value=studentMark[i][3]
                    sheet.cell(row=i+1,column=5).value=studentMark[i][4]
                    sheet.cell(row=i+1,column=6).value=studentMark[i][5]
                    sheet.cell(row=i+1,column=7).value=studentMark[i][6]
                    sheet.cell(row=i+1,column=8).value=studentMark[i][7]
                    sheet.cell(row=i+1,column=9).value=studentMark[i][8]
                except:
                    flash("Internal Marks not entered for "+cls)
            wb.save(path)


@app.route('/promoteSem')
def promoteSem():
    classes=existingClasses()
    promoter(classes)
    return render_template("essentials.html",classes=classes,promote="Promote Sem")


@app.route('/promoteYear')
def promoteYear():
    classes=existingClasses()
    promoter(classes)
    return render_template("essentials.html",classes=classes,promote="Promote Year")


@app.route('/deleteData',methods=['POST'])
def deleteData():
    if request.method=='POST':
        classes=existingClasses()
        marks=[]
        for i in classes:
            temp=os.listdir("/home/JAKdevPortal/mysite/static/internalmark/"+i)
            materials=os.listdir("/home/JAKdevPortal/mysite/static/materials/"+i)
            for j in temp:
                os.remove("/home/JAKdevPortal/mysite/static/internalmark/"+i+"/"+j)
            for j in materials:
                shutil.rmtree("/home/JAKdevPortal/mysite/static/materials/"+i+"/"+j)
            shutil.copy("/home/JAKdevPortal/mysite/static/namelist/"+i+".xlsx","/home/JAKdevPortal/mysite/static/internalmark/"+i+".xlsx")
        cursor.execute("update general set staffs='',subjects=''")
        paths=["/home/JAKdevPortal/mysite/static/{}Internal Assessment I.xlsx","/home/JAKdevPortal/mysite/static/{}Internal Assessment II.xlsx","/home/JAKdevPortal/mysite/static/{}Model Exam.xlsx","/home/JAKdevPortal/mysite/static/attendance/{}.xlsx"]
        for cls in classes:
            for path in paths:
                os.remove(path.format(cls))
        for cls in classes:
            #copy only column 1,2,4 from namlist/cls.xlsx to attendance/cls.xlsx
            wb=load_workbook("/home/JAKdevPortal/mysite/static/namelist/"+cls+".xlsx")
            sheet_obj = wb.active
            m_row = sheet_obj.max_row
            m_col=sheet_obj.max_column
            wb1=Workbook()
            sheet=wb1.active
            sheet.cell(row=1,column=1).value="REG NO"
            sheet.cell(row=1,column=2).value="NAME"
            sheet.cell(row=1,column=3).value="PHONE NO"
            for i in range(2, m_row + 1):
                for j in range(1,m_col+1):
                    cell_obj = sheet_obj.cell(row = i, column = j)
                    sheet.cell(row=i,column=j).value=cell_obj.value
            wb1.save("/home/JAKdevPortal/mysite/static/attendance/"+cls+".xlsx")

        cursor.execute("update staff set subjects='',classes='',code=''")
        cursor.execute("update mainmaterials set unit1='',unit2='',unit3='',unit4='',unit5=''")
        cursor.execute("update additionalmaterials set unit1='',unit2='',unit3='',unit4='',unit5=''")
        sqliteConnection.commit()
        if request.form["clicked"]=="Promote Sem":
            flash("Semester promoted successfully")
            return redirect(url_for('createClass'))
        elif request.form["clicked"]=="Promote Year":
            cursor.execute("alter table a3 rename to A4")
            cursor.execute("CREATE TABLE A3(rollno varchar(12),password varchar(22),name varchar(50),boarding varchar(3),cgpa varchar(4),email varchar(40),phoneno varchar(11),certifications varchar(250),projects varchar(200),prizes varchar(250),internships varchar(300),parentNo varchar(12),linkedin varchar(100),cgpa varchar(55),sex varchar(2));")
            sqliteConnection.commit()
            data=cursor.execute("select * from a2").fetchall()
            for i in data:
                cursor.execute("insert into A3 values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",i)
            data=cursor.execute("select * from b2").fetchall()
            for i in data:
                cursor.execute("insert into A2 values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",i)
            sqliteConnection.commit()
        return redirect(url_for('createClass'))

@app.route('/deleteClass',methods=['POST'])
def deleteClass():
    if request.method=='POST':
        delClass=request.form['clicked'][-2:]
        flash('Class deleted successfully')
        return redirect(url_for('createClass'))


@app.route('/staffDetails')
def staffDetails():
    cursor.execute("select Name,Email,Designation,Qualification,phoneno,dept from staff")
    staffs=cursor.fetchall()
    return render_template('staffDetails.html',staffs=staffs)

@app.route('/newStaff',methods=['POST'])
def newStaff():
    if request.method=='POST':
        facultyName=request.form['facultyName']
        deptStaff=request.form.get("deptStaff")
        os.mkdir("/home/JAKdevPortal/mysite/static/papers/"+facultyName)
        if request.form['clicked']=='Add Faculty':
            cursor.execute("insert into staff values(?,?,?,?,?,?,?,?,?,?,?,?)",(facultyName,facultyName,facultyName,"","","","","","","","",deptStaff))
            sqliteConnection.commit()
            flash("Staff added successfully")
            return redirect(url_for('staffDetails'))

@app.route('/removeStaff/<staffName>')
def removeStaff(staffName):
    cursor.execute("delete from staff where name=?",(staffName,))

    sqliteConnection.commit()
    os.rmdir("/home/JAKdevPortal/mysite/static/papers/"+staffName)
    staffName=staffName.replace(" ","")
    profilesList=os.listdir("/home/JAKdevPortal/mysite/static/profiles/STAFFS/")
    for i in profilesList:
        if staffName in i:
            os.remove("/home/JAKdevPortal/mysite/static/profiles/STAFFS/"+i)
    flash("Staff removed successfully")
    return redirect(url_for('staffDetails'))

@app.route('/assignCounsellors')
def assignCounsellors():
    staffs=selectQueryHelp(cursor.execute("select name from staff").fetchall())
    counsellors=cursor.execute("select * from counsellors").fetchall()
    return render_template('assignCounsellors.html',staffs=staffs,counsellors=counsellors)

@app.route('/newCounsellor',methods=['POST'])
def newCounsellor():
    if request.method=='POST':
        fromRollNo=request.form['fromRollNo']
        toRollNo=request.form['toRollNo']
        counsellor=request.form.get('counsellor')
        if request.form['clicked']=='Assign Counsellor':
            cursor.execute("insert into counsellors values(?,?,?)",(fromRollNo,toRollNo,counsellor))
            sqliteConnection.commit()
            flash('Counsellor assigned successfully')
            return redirect(url_for('assignCounsellors'))

@app.route('/deleteCounsellor/<fromRollNo>/<toRollNo>/<counsellor>')
def deleteCounsellor(fromRollNo,toRollNo,counsellor):
    cursor.execute("delete from counsellors where start=? and end=? and counsellor=?",(fromRollNo,toRollNo,counsellor))
    sqliteConnection.commit()
    flash('Counsellor deleted successfully')
    return redirect(url_for('assignCounsellors'))

@app.route('/assignSubjects')
def assignSubjects():
    cursor.execute("select class from general where subjects=''")
    classes=selectQueryHelp(cursor.fetchall())
    classes=[clsToYear[i] for i in classes]
    staffs=selectQueryHelp(cursor.execute("select name from staff").fetchall())
    return render_template('assignSubjects.html',classes=classes,staffs=staffs)


@app.route('/newSubjects', methods=['POST'])
def newSubjects():
    if request.method=='POST':
        subjectNameString=""
        subjectTeacherString=""
        year=request.form.get('class')
        yr=yearToCls[year]
        mainList=[]
        for i in range(1,7):
            try:
                subjectName=request.form['subjectName'+str(i)]
                subjectTeacher=request.form.get('staff'+str(i))
                subjectCode=request.form['subjectCode'+str(i)]
                if subjectName!="" and subjectCode!="":
                    existOrNot=cursor.execute("select * from mainmaterials where subject=? and staff=?", (subjectName,subjectTeacher)).fetchall()
                    tempList=[subjectName,subjectCode,subjectTeacher]

                    if existOrNot==None or existOrNot==() or existOrNot==[]:
                        cursor.execute("insert into mainmaterials values(?,?,?,?,?,?,?)",(subjectName,"","","","","",subjectTeacher))
                        cursor.execute("insert into additionalmaterials values(?,?,?,?,?,?,?)",(subjectName,"","","","","",subjectTeacher))
                        sqliteConnection.commit()
                    subjectNameString+=tempList[0]+","
                    os.mkdir("/home/JAKdevPortal/mysite/static/materials/"+yr+"/"+tempList[0])
                    shutil.copy("/home/JAKdevPortal/mysite/static/internalmark/"+yr+".xlsx","/home/JAKdevPortal/mysite/static/internalmark/"+yr+"/"+tempList[0]+".xlsx")
                    for j in range(1,6):
                        os.mkdir("/home/JAKdevPortal/mysite/static/materials/"+yr+"/"+tempList[0]+"/"+"unit"+str(j))
                        os.mkdir("/home/JAKdevPortal/mysite/static/materials/"+yr+"/"+tempList[0]+"/"+"unit"+str(j)+"/"+"additional materials")
                    subjectTeacherString+=tempList[2]+","
                    mainList.append(tempList)
            except:
                pass
        os.remove("/home/JAKdevPortal/mysite/static/internalmark/"+yr+".xlsx")
        cursor.execute("update general set subjects=?,staffs=? where class=?",(subjectNameString,subjectTeacherString,yr))
        sqliteConnection.commit()
        for i in mainList:
            cursor.execute("select subjects,code,classes from staff where name=?",(i[2],))
            temp=cursor.fetchone()
            cursor.execute("update staff set subjects=?,code=?,classes=? where name=?",(temp[0]+i[0]+",",temp[1]+i[1]+",",temp[2]+yr+",",i[2]))
            sqliteConnection.commit()
        flash('Subjects added successfully...')
        return redirect(url_for('assignSubjects'))

@app.route("/adminSettings")
def adminSettings():
    return render_template('adminSettings.html')

@app.route("/updateAdminProfile",methods=['POST'])
def updateAdminProfile():
    if request.method=="POST":
        userID=request.form['userID']
        password=request.form['password']
        if userID!="":
            cursor.execute("update admin set userID=?",(userID,))
        if password!="":
            cursor.execute("update admin set password=?",(password,))
        if request.form["clicked"]=="Update":
            sqliteConnection.commit()
            flash('Password successfully changed')
            return redirect(url_for('adminSettings'))

def calculateLeavePercentage(rollno,cls):
    path="/home/JAKdevPortal/mysite/static/attendance/"+cls+".xlsx"
    wb=load_workbook(path)
    ws=wb.active
    totalDays=ws.max_column-3
    for i in range(1,ws.max_row+1):
        if str(ws.cell(row=i,column=1).value)==rollno:
            presentDays=0
            for j in range(4,ws.max_column+1):
                if str(ws.cell(row=i,column=j).value).lower()!='ab' and str(ws.cell(row=i,column=j).value).lower()!='a':
                    presentDays+=1
            if presentDays==0:
                return totalDays,0
            return (totalDays-presentDays), round((presentDays/totalDays)*100,2)


@app.route("/adminInternal")
def adminInternal():
    Classes=existingClasses()
    try:
        subjects=cursor.execute("select subjects from general where class = ?",(session["class"],)).fetchone()[0].split(',')[:-1]
    except:
        subjects=[]
    return render_template ("adminInternal.html",Classes=Classes,subjects=subjects,selectedClass="nil",selectedTest="nil")

@app.route("/internalMarksAdmin",methods=['POST'])
def internalMarksAdmin():
    if request.method=="POST":
        selectedClass=request.form.get("class")
        cls=yearToCls[selectedClass]
        selectedTest=request.form.get("test")
        session["selectedClass"]=cls
        session["selectedTest"]=selectedTest
        subjects=cursor.execute("select subjects from general where class = ?",(cls,)).fetchone()[0].split(',')[:-1]
        studentMark=[]
        path="/home/JAKdevPortal/mysite/static/namelist/"+cls+".xlsx"
        wb=load_workbook(path)
        ws=wb.active
        for i in range(2,ws.max_row+1):
            studentMark.append([ws.cell(row=i,column=1).value,ws.cell(row=i,column=2).value,cursor.execute("select parentNo from {} where rollno=?".format(cls),(ws.cell(row=i,column=1).value,)).fetchone()[0]])

        path="/home/JAKdevPortal/mysite/static/internalmark/"+cls+"/"
        for subject in subjects:
            wb=load_workbook(path+subject+".xlsx")
            ws=wb.active
            flag=1
            for i in range(4,ws.max_column+1):
                if ws.cell(row=1,column=i).value==selectedTest:
                    flag=0
                    for j in range(2,ws.max_row+1):
                        studentMark[j-2].append(ws.cell(row=j,column=i).value)
                    break
            if flag:
                for i in studentMark:
                    i.append("nil")
        session["whatsapp"]=studentMark
        #write student mark in the following xlsx file
        path="/home/JAKdevPortal/mysite/static/internalmark/"+"admin"+cls+".xlsx"
        wb=Workbook()
        sheet=wb.active
        sheet.cell(row=1,column=1).value="REG NO"
        sheet.cell(row=1,column=2).value="NAME"
        for subject in range(len(subjects)):
            sheet.cell(row=1,column=3+subject).value=subjects[subject]
        for i in range(1,len(studentMark)+1):
            sheet.cell(row=i+1,column=1).value=studentMark[i-1][0]
            sheet.cell(row=i+1,column=2).value=studentMark[i-1][1]
            #subjects Marks
            try:
                sheet.cell(row=i+1,column=3).value=studentMark[i-1][3]
                sheet.cell(row=i+1,column=4).value=studentMark[i-1][4]
                sheet.cell(row=i+1,column=5).value=studentMark[i-1][5]
                sheet.cell(row=i+1,column=6).value=studentMark[i-1][6]
                sheet.cell(row=i+1,column=7).value=studentMark[i-1][7]
                sheet.cell(row=i+1,column=8).value=studentMark[i-1][8]
            except:
                flash("Internal Marks not entered for "+cls)
        wb.save(path)
        return render_template("adminInternal.html",studentMark=studentMark,subjects=subjects,selectedClass=yearToCls[selectedClass],selectedTest=selectedTest,Classes=existingClasses())

def internalMarksAdmin(cls,selectedTest):
    subjects=cursor.execute("select subjects from general where class = ?",(cls,)).fetchone()[0].split(',')[:-1]
    studentMark=[]
    path="/home/JAKdevPortal/mysite/static/namelist/"+cls+".xlsx"
    wb=load_workbook(path)
    ws=wb.active
    for i in range(2,ws.max_row+1):
        studentMark.append([ws.cell(row=i,column=1).value,ws.cell(row=i,column=2).value,ws.cell(row=i,column=4).value,])
    path="/home/JAKdevPortal/mysite/static/internalmark/"+cls+"/"

    for subject in subjects:
        wb=load_workbook(path+subject+".xlsx")
        ws=wb.active
        flag=1
        for i in range(4,ws.max_column+1):
            if ws.cell(row=1,column=i).value==selectedTest:
                flag=0
                for j in range(2,ws.max_row+1):
                    studentMark[j-2].append(ws.cell(row=j,column=i).value)
                break
        if flag:
            for i in range(len(studentMark)):
                studentMark[i].append("nil")
    studentMark.insert(0,subjects)
    return studentMark

@app.route("/studentLogin",methods=['POST'])
def studentLogin():
    if request.method=='POST':
        username = request.form['userID']
        password = request.form['password']
        classes=existingClasses()
        for cls in classes:
            cursor.execute("SELECT * FROM {} WHERE rollno = ? AND password = ?".format(cls), (username, password))
            data=cursor.fetchone()
            if data:
                session['logged_in'] = True
                session["profile"]=""
                profiles=os.listdir("/home/JAKdevPortal/mysite/static/profiles/"+cls)
                for i in profiles:
                    if i.startswith(str(data[0])):
                        session["profile"]=i
                session["data"] = data
                session["class"] = cls
                session["name"]=data[2]
                session["rollno"]=int(data[0])
                session["counsellor"]="nil"
                session["leavePercentage"]=calculateLeavePercentage(data[0],cls)
                #finding counsellor
                try:
                    cursor.execute("select counsellor from counsellors where ? between start and end",(session["rollno"],))
                    session["counsellor"]=cursor.fetchone()[0]
                except:
                    pass
                return redirect(url_for('profile'))
        else:
            flash("Incorrect password...")
            return redirect(url_for('index'))




@app.route('/staffLogin', methods=['POST'])
def staffLogin():
    if request.method=='POST':
        username = request.form['userID']
        password = request.form['password']
        cursor.execute("select * from admin where userid=? and password=?",(username,password))
        data=cursor.fetchone()
        if data:
            session['logged_in'] = True
            session['data']=data
            session["attendanceReportPath"]="None"

            return redirect(url_for('createClass'))
        cursor.execute("SELECT * FROM staff WHERE userid = ? AND password = ?", (username, password))
        data = cursor.fetchone()
        if data:
            session['logged_in'] = True
            session["profile"]=""

            profiles=os.listdir("/home/JAKdevPortal/mysite/static/profiles/STAFFS/")
            for i in profiles:
                firstName=data[2].replace(" ","")
                if i.startswith(firstName):
                    session["profile"]=i
            session["data"] = data
            return redirect(url_for('staffProfile'),)
        flash('Incorrect password...')
        return redirect(url_for('index'))

@app.route('/leaveRequest', methods=['POST'])
def leaveRequest():
    if request.method=='POST':
        leaveType=request.form.get('type')
        reason=request.form['reason']
        fromDate=request.form['fromDate']
        toDate=request.form['toDate']
        currentDate=datetime.now().strftime("%Y-%d-%m")
        currentTime=datetime.now().strftime("%H:%M:%S")
        cursor.execute("insert into leave values(?,?,?,?,?,?,?,?,?,?)",(session["name"]+"("+str(session["rollno"])+")",currentDate,currentTime,fromDate,toDate,leaveType,reason,"pending",session["counsellor"],session["class"]))
        sqliteConnection.commit()
        flash("Request sent successfully...")
        return redirect(url_for('leaveApplication'))

@app.route('/leaveStatus', methods=['POST'])
def leaveStatus():
    if request.method=='POST':
        status,applicant,fromdate=request.form['clicked'].split(';')
        cursor.execute("update leave set status = ? where appliedBy = ? and fromDate = ?",(status+"ed",applicant,fromdate))
        sqliteConnection.commit()
        return redirect(url_for('staffLeaveLetters'))

@app.route('/addCertificate', methods=['POST'])
def addCertificate():
    if request.method=='POST':
        certificateName=request.form['certificateName']
        certificateFile=request.files['certificateFile']
        extension=certificateFile.filename.split('.')[-1]
        if request.form["clicked"]=="Add":
            existingCertificates=cursor.execute("select certifications from {} where rollno = ?".format(session["class"]),(session["rollno"],)).fetchone()[0]
            existingCertificates=existingCertificates+certificateName+"."+extension+"%"
            cursor.execute("update {} set certifications = ? where rollno = ?".format(session["class"]),(existingCertificates,session["rollno"]))
            sqliteConnection.commit()
            certificateFile.save("/home/JAKdevPortal/mysite/static/certificates/"+session["class"]+"/"+str(session["rollno"])+"/"+certificateName+"."+extension)
            flash("Uploaded successfully...")
            return redirect(url_for('certifications'))

@app.route('/deleteCertificate/<location>/<certificateName>')
def deleteCertificate(location,certificateName):
    location=location.replace(',',"/")
    os.remove(location+"/"+certificateName)
    existingClasses=cursor.execute("select certifications from {} where rollno = ?".format(session["class"]),(session["rollno"],)).fetchone()[0].split("%")
    existingClasses.remove(certificateName)
    existingClasses="%".join(existingClasses)
    cursor.execute("update {} set certifications = ? where rollno = ?".format(session["class"]),(existingClasses,session["rollno"]))
    sqliteConnection.commit()

    return redirect(url_for('certifications'))

@app.route("/addInternship", methods=["POST"])
def addInternship():
    if request.method=='POST':
        internCompany=request.form['internCompany']
        internDomain=request.form['internDomain']
        internDuration=request.form['internDuration']
        internFile=request.files['internFile']
        extension=internFile.filename.split('.')[-1]
        if request.form["clicked"]=="Add":
            existingInternships=cursor.execute("select internships from {} where rollno = ?".format(session["class"]),(session["rollno"],)).fetchone()[0]
            existingInternships+=internCompany+"."+internDomain+"."+internDuration+"."+extension+"%"
            cursor.execute("update {} set internships = ? where rollno = ?".format(session["class"]),(existingInternships,session["rollno"]))
            sqliteConnection.commit()
            internFile.save("/home/JAKdevPortal/mysite/static/internships/"+session["class"]+"/"+str(session["rollno"])+"/"+internCompany+"."+internDomain+"."+internDuration+"."+extension)
            flash('Internship added successfully!!...')
            return redirect(url_for('internships'))

@app.route("/deleteInternship/<location>/<internshipName>")
def deleteInternship(location,internshipName):
    location=location.replace(',',"/")
    os.remove(location+"/"+internshipName)
    existingInternships=cursor.execute("select internships from {} where rollno = ?".format(session["class"]),(session["rollno"],)).fetchone()[0].split("%")
    existingInternships.remove(internshipName)
    existingInternships="%".join(existingInternships)
    cursor.execute("update {} set internships = ? where rollno = ?".format(session["class"]),(existingInternships,session["rollno"]))
    sqliteConnection.commit()
    return redirect(url_for('internships'))


@app.route('/addPrize', methods=['POST'])
def addPrize():
    if request.method=='POST':
        prizeName=request.form['prizeName']
        competitionName=request.form['competitionName']
        prize=request.form['prize']
        prizeFile=request.files['prizeFile']
        extension=prizeFile.filename.split('.')[-1]
        if request.form["clicked"]=="Add":
            existingPrizes=cursor.execute("select prizes from {} where rollno = ?".format(session["class"]),(session["rollno"],)).fetchone()[0]
            existingPrizes+=prizeName+"."+competitionName+"."+prize+"."+extension+"%"

            cursor.execute("update {} set prizes = ? where rollno = ?".format(session["class"]),(existingPrizes,session["rollno"]))
            sqliteConnection.commit()
            prizeFile.save("/home/JAKdevPortal/mysite/static/prizes/"+session["class"]+"/"+str(session["rollno"])+"/"+prizeName+"."+competitionName+"."+prize+"."+extension)
            return redirect(url_for('prizes'))

@app.route("/deletePrize/<location>/<prizeName>")
def deletePrize(location,prizeName):
    location=location.replace(',',"/")
    os.remove(location+"/"+prizeName)
    existingPrizes=cursor.execute("select prizes from {} where rollno = ?".format(session["class"]),(session["rollno"],)).fetchone()[0].split("%")
    existingPrizes.remove(prizeName)
    existingPrizes="%".join(existingPrizes)
    cursor.execute("update {} set prizes = ? where rollno = ?".format(session["class"]),(existingPrizes,session["rollno"]))
    sqliteConnection.commit()
    return redirect(url_for('prizes'))



@app.route('/addProject', methods=['POST'])
def addProject():
    if request.method=='POST':
        projectName=request.form['projectName']
        projectFile=request.files['projectFile']
        extension=projectFile.filename.split('.')[-1]
        if request.form["clicked"]=="Add":
            existingProjects=cursor.execute("select projects from {} where rollno = ?".format(session["class"]),(session["rollno"],)).fetchone()[0]
            existingProjects+=projectName+"."+extension+"%"
            cursor.execute("update {} set projects = ? where rollno = ?".format(session["class"]),(existingProjects,session["rollno"]))
            sqliteConnection.commit()
            projectFile.save("/home/JAKdevPortal/mysite/static/projects/"+session["class"]+"/"+str(session["rollno"])+"/"+projectName+"."+extension)
            flash('Project added Successfully!!...')
            return redirect(url_for('projects'))

@app.route("/deleteProject/<location>/<projectName>")
def deleteProject(location,projectName):
    location=location.replace(',',"/")
    os.remove(location+"/"+projectName)
    existingProjects=cursor.execute("select projects from {} where rollno = ?".format(session["class"]),(session["rollno"],)).fetchone()[0].split("%")
    existingProjects.remove(projectName)
    existingProjects="%".join(existingProjects)
    cursor.execute("update {} set projects = ? where rollno = ?".format(session["class"]),(existingProjects,session["rollno"]))
    sqliteConnection.commit()
    return redirect(url_for('projects'))

@app.route('/staffs')
def staffs():
    cursor.execute("select name,email,phoneno from staff")
    staffs=cursor.fetchall()
    profilePic=[]
    available=os.listdir("/home/JAKdevPortal/mysite/static/profiles/STAFFS/")
    for i in staffs:
        name=i[0].replace(" ","")
        #startswith
        for j in available:
            if j.startswith(name):
                profilePic.append(j)
                break
    return render_template('staffs.html',staffs=staffs,profilePic=profilePic)

def enumerate_subjects(subjects):
    return enumerate(subjects, start=1)
app.jinja_env.filters['enumerate_subjects'] = enumerate_subjects

@app.route("/addMaterial", methods=["POST"])
def addMaterial():
    if request.method=='POST':
        subject=request.form['subject']
        unit=request.form['unit']
        materialType=request.form['materialType']
        materialName=request.form['materialName']
        materialFile=request.files['materialFile']
        extension=materialFile.filename.split('.')[-1]
        if request.form["clicked"]=="Upload":
            classes=cursor.execute("select class from general where subjects like '%"+subject+"%'").fetchall()
            classes=selectQueryHelp(classes)
            print(classes)
            if materialType=="Main Material":
                existingMaterial=os.listdir("/home/JAKdevPortal/mysite/static/materials/"+classes[0]+"/"+subject+"/"+unit+"/")
                if len(existingMaterial)!=1:
                    indexx=existingMaterial.index("additional materials")
                    if indexx==0:
                        indexx=1
                    else:
                        indexx=0
                for cls in classes:
                    try:
                        os.remove("/home/JAKdevPortal/mysite/static/materials/"+cls+"/"+subject+"/"+unit+"/"+existingMaterial[indexx])
                    except:
                        pass
                materialFile.save("/home/JAKdevPortal/mysite/static/materials/"+classes[0]+"/"+subject+"/"+unit+"/"+materialName+"."+extension)
                for cls in classes[1:]:
                    shutil.copy("/home/JAKdevPortal/mysite/static/materials/"+classes[0]+"/"+subject+"/"+unit+"/"+materialName+"."+extension,"/home/JAKdevPortal/mysite/static/materials/"+cls+"/"+subject+"/"+unit+"/"+materialName+"."+extension)
                cursor.execute("update mainmaterials set {} = ? where subject = ? ".format(unit),(materialName+"."+extension,subject))
            else:
                materialFile.save("/home/JAKdevPortal/mysite/static/materials/"+classes[0]+"/"+subject+"/"+unit+"/"+"additional materials"+"/"+materialName+"."+extension)
                for cls in classes[1:]:
                    shutil.copy("/home/JAKdevPortal/mysite/static/materials/"+classes[0]+"/"+subject+"/"+unit+"/"+"additional materials"+"/"+materialName+"."+extension,"/home/JAKdevPortal/mysite/static/materials/"+cls+"/"+subject+"/"+unit+"/"+"additional materials"+"/"+materialName+"."+extension)

                existingMaterial=cursor.execute("select {} from additionalmaterials where subject = ?".format(unit),(subject,)).fetchone()[0]
                existingMaterial+=(materialName+"."+extension+"%")
                cursor.execute("update additionalmaterials set {} = ? where subject = ?".format(unit),(existingMaterial,subject))
            sqliteConnection.commit()
            flash('Added Successfully...')
            return redirect(url_for('staffMaterials'))

@app.route("/deleteMaterial",methods=["POST"])
def deleteMaterial():
    if request.method=='POST':
        subject=request.form.get("subject")
        unit=request.form['unit'].lower()
        materialType=request.form['materialType']
        session["selectedDelete"]=[subject,unit,materialType]
        classes=cursor.execute("select class from general where subjects like '%"+subject+"%'").fetchall()
        classes=selectQueryHelp(classes)
        session["deleteClasses"]=classes
        session["deleteDetails"]=[subject,unit,materialType]
        if materialType=="Main Material":
            materials=os.listdir("/home/JAKdevPortal/mysite/static/materials/"+classes[0]+"/"+subject+"/"+unit.lower()+"/")
            materials.remove('additional materials')
        else:
            materials=os.listdir("/home/JAKdevPortal/mysite/static/materials/"+classes[0]+"/"+subject+"/"+unit.lower()+"/"+"additional materials/")
        if request.form["clicked"]=="Show":
            session["deleteMaterial"]=materials
            return redirect(url_for('staffMaterials'))


@app.route("/deleteMaterial/<material>")
def deleteMaterials(material):
    classes=session["deleteClasses"]
    deleteDetails=session["deleteDetails"]
    if deleteDetails[2]=="Main Material":
        for cls in classes:
            path="/home/JAKdevPortal/mysite/static/materials/"+cls+"/"+deleteDetails[0]+"/"+deleteDetails[1]+"/"
            try:
                os.remove(path+material)
                cursor.execute("update mainmaterials set {} ='' where subject=?".format(deleteDetails[1]),(deleteDetails[0],))
                sqliteConnection.commit()
            except:
                pass
    else:
        for cls in classes:
            path="/home/JAKdevPortal/mysite/static/materials/"+cls+"/"+deleteDetails[0]+"/"+deleteDetails[1]+"/"
            try:
                os.remove(path+"additional materials/"+material)
                temp=cursor.execute("select {} from additionalMaterials where subject=?".format(deleteDetails[1]),(deleteDetails[0],)).fetchone()[0].split("%")
                if len(temp)==1:
                    m=""
                else:
                    try:
                        temp.remove(material)
                        m=""
                        for i in temp:
                            if i!="":
                                m+=i
                                m+="%"
                        print(m)
                    except:
                        pass
                cursor.execute("update additionalmaterials set {} =? where subject=?".format(deleteDetails[1]),(m,deleteDetails[0]))
                sqliteConnection.commit()
            except:
                pass
    session.pop("deleteDetails")
    session["deleteMaterial"]=""
    session["selectedDelete"]=""
    return redirect(url_for('staffMaterials'))


if __name__ == '__main__':
    try:
        app.secret_key="secret_key"
        app.run(debug=True)
    except:
        pass
